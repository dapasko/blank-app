import streamlit as st
import pandas as pd
import plotly.express as px
import io
from openpyxl.utils import get_column_letter
import json
import os
from datetime import datetime
from typing import List
def hash_dataframe(df):
    """Хэширует DataFrame для поддержки @st.cache_data"""
    return pd.util.hash_pandas_object(df).sum()

# предполагаем, что mapping.json лежит рядом с этим файлом
here = os.path.dirname(__file__)
with open(os.path.join(here, "mapping.json"), "r", encoding="utf-8") as f:
    VARIANT_TO_SYSTEM = json.load(f)

CHANNEL_MAPPING = {
    "Входящие звонки": ["Входящие звонки"],
    "Оффлайн функционал": ["Работа", "Дежурство", "Письма"],
    "Отработки-Доп рабочие интервалы": ["Доп рабочие интервалы", "Отработка"],
    "Чаты": ["Чат"],
    "Перерывы": ["Свободное время", "Обед", "Перерыв"],
    "Отпуск": ["Отпуск", "Учебный отпуск"],
    "Больничный" : ["Больничный"]
}

def inject_custom_css():
    """
    Внедряет пользовательские стили в приложение Streamlit.
    Улучшает внешний вид интерфейса с помощью CSS.
    Использует цвета из текущей темы Streamlit (light/dark).
    """
    # Получаем тип темы: 'light' или 'dark'
    theme_type = st.context.theme.type

    # Определяем цвета в зависимости от темы
    if theme_type == "light":
        background = "#FFFFFF"     # белый фон
        text_color = "#000000"     # чёрный текст
    else:
        background = "#1E1E1E"     # тёмный фон
        text_color = "#FFFFFF"     # белый текст

    primary_color = "#3498db"      # основной цвет (можно оставить фиксированным)

    # Внедряем CSS
    st.markdown(f"""
        <style>
            :root {{
                --bg-color: {background};
                --text-color: {text_color};
                --primary-color: {primary_color};
            }}

            /* Анимации */
            @keyframes fadeIn {{
                0% {{ opacity: 0; transform: translateY(10px); }}
                100% {{ opacity: 1; transform: translateY(0); }}
            }}

            /* Основные элементы */
            .stApp {{
                background: var(--bg-color);
                color: var(--text-color);
                animation: fadeIn 0.5s ease-in;
            }}

            /* Заголовки */
            h1 {{
                color: var(--primary-color);
                border-bottom: 3px solid var(--primary-color);
                padding-bottom: 0.5rem;
                margin-bottom: 1.5rem !important;
            }}

            /* Карточки фильтров */
            .filter-card {{
                background: white;
                border-radius: 12px;
                padding: 1.5rem;
                box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);
                margin-bottom: 1.5rem;
            }}

            /* Кнопки */
            .stButton>button {{
                transition: all 0.3s ease;
                border: 2px solid var(--primary-color) !important;
            }}
            .stButton>button:hover {{
                transform: translateY(-2px);
                box-shadow: 0 7px 14px rgba(52,152,219,0.2);
            }}

            /* Иконки */
            .material-icons {{
                vertical-align: middle;
                margin-right: 5px;
            }}
        </style>
    """, unsafe_allow_html=True)

def filter_data(df_act, selected_variants, selected_main_acts):
    """
    Фильтрует DataFrame активности по выбранным группам навыков и типам активности.

    Args:
        df_act (pd.DataFrame): Исходный DataFrame активности
        selected_variants (list[str]): Список выбранных групп навыков
        selected_main_acts (list[str]): Список типов активности для фильтрации

    Returns:
        pd.DataFrame: Отфильтрованный DataFrame
    """
    return df_act[
        df_act['skill_variant'].isin(selected_variants) &
        df_act['main_act'].isin(selected_main_acts)
    ]

def process_forecast(forecast_file, selected_system_groups, selected_channels, selected_forecast_col):
    """
       Обрабатывает данные прогноза: загружает, проверяет, фильтрует и возвращает df_forecast

    Args:
        forecast_file (io.BytesIO): Загруженный файл прогноза
        selected_system_groups (list[str]): Системные группы, по которым фильтруется прогноз
        selected_channels (list[str]): Каналы коммуникации для фильтрации
        selected_forecast_col (str): Название колонки прогноза (например, 'Прогноз Raw')

    Returns:
        pd.DataFrame: Обработанный DataFrame прогноза
    """
    df_forecast = load_forecast(forecast_file)

    # --- Проверка: файл прогноза загружен, но пустой ---
    if forecast_file is not None and df_forecast.empty:
        st.warning("⚠️ Файл прогноза загружен, но он пустой или содержит некорректные данные.")
        st.markdown("Проверьте формат и содержание файла прогноза.")

    # --- Проверка: прогноз не загружен, но пользователь выбрал колонку прогноза ---
    if forecast_file is None and selected_forecast_col:
        st.info("ℹ️ Файл прогноза не загружен. Прогноз не будет отображён.")
        st.markdown("Загрузите файл прогноза в боковой панели, чтобы использовать прогнозные значения.")

    if not df_forecast.empty:
        df_forecast = df_forecast[
            df_forecast['system_group'].isin(selected_system_groups) &
            df_forecast['Канал коммуникации'].isin(selected_channels)
        ]

        # Выбираем только нужную колонку прогноза
        df_forecast = df_forecast[[
            'Дата', 'Время', 'system_group', 'Канал коммуникации', selected_forecast_col
        ]].rename(columns={selected_forecast_col: 'Прогноз'})

    return df_forecast

def render_filters(df_act):
    """
    Отображает фильтры в боковой панели и возвращает выбранные параметры.

    Позволяет пользователю выбрать:
    - Колонку прогноза
    - Группы навыков
    - Каналы коммуникации

    Args:
        df_act (pd.DataFrame): DataFrame активности для получения уникальных групп

    Returns:
        dict: Словарь с фильтрами:
            - selected_forecast_col (str): Выбранная колонка прогноза
            - selected_variants (list[str]): Выбранные группы навыков
            - selected_channels (list[str]): Выбранные каналы коммуникации
            - selected_main_acts (list[str]): Активности, соответствующие каналам
    """
    with st.sidebar.expander("🔮 **Параметры фильтрации**", expanded=True):
        # Фильтр версии прогноза
        forecast_columns = ['Прогноз Raw', 'Прогноз Abs_new', 'Прогноз Full']
        selected_forecast_col = st.selectbox(
            "Выберите колонку прогноза",
            options=forecast_columns,
            index=0,
            key="forecast_col_filter"
        )

        # Фильтр скилл-групп
        variants = sorted(df_act['skill_variant'].dropna().astype(str).unique())
        selected_variants = st.multiselect(
            "**Группы навыков**",
            options=variants,
            default=variants,
            help="Выберите одну или несколько групп",
            placeholder="Поиск...",
            key="skill_filter",
            format_func=lambda x: f"🎯 {x}"
        )

        # Индикатор выбора
        st.write(
            f"<span style='font-size:0.9em; color:#7f8c8d;'>Выбрано: {len(selected_variants)} групп</span>",
            unsafe_allow_html=True
        )

        # Разделитель
        st.markdown("<hr style='margin:1.5rem 0; border-color:#eee;'>", unsafe_allow_html=True)

        # Фильтр каналов
        selected_channels = st.multiselect(
            "**Каналы коммуникации**",
            options=list(CHANNEL_MAPPING.keys()),
            default=["Входящие звонки", "Чаты"],
            format_func=lambda x: {
                "Входящие звонки": "📞 Входящие звонки",
                "Оффлайн функционал": "💻 Оффлайн",
                "Чаты": "💬 Чаты",
                "Перерывы": "☕ Перерывы",
                "Отработки-Доп рабочие интервалы": "👨‍💻 Отработка-Допка",
                "Отпуск": "🏖️ Отпуск",
                "Больничный": "🏥 Больничный"
            }[x],
            key="channel_filter"
        )

        selected_main_acts = [act for channel in selected_channels for act in CHANNEL_MAPPING[channel]]

        # Кнопка применения
        if st.button(
                "🔄 Обновить данные",
                type="primary",
                use_container_width=True,
                help="Применить выбранные фильтры"
        ):
            st.rerun()

    return {
        'selected_forecast_col': selected_forecast_col,
        'selected_variants': selected_variants,
        'selected_channels': selected_channels,
        'selected_main_acts': selected_main_acts
    }


def filter_by_period(df_filtered, start_dt, end_dt):
    """
    Фильтрует DataFrame активности по временному диапазону.

    Args:
        df_filtered (pd.DataFrame): Отфильтрованный DataFrame активности
        start_dt (datetime): Начало временного диапазона
        end_dt (datetime): Конец временного диапазона

    Returns:
        pd.DataFrame: Активность в рамках выбранного периода
    """
    if start_dt is None or end_dt is None:
        return df_filtered
    return df_filtered[
        (df_filtered['start'] < end_dt) & (df_filtered['end'] > start_dt)
        ]

def calculate_plan(df_period, times: List[datetime], is_hourly=True):
    """
    Рассчитывает план по заданным временным слотам.

    Args:
        df_period (pd.DataFrame): DataFrame активности в рамках периода
        times (list[datetime]): Список временных слотов
        is_hourly (bool): Режим — по часам (True) или по дням (False)

    Returns:
        pd.DataFrame: DataFrame с колонкой 'План' (в часах)
    """
    plan_list = []
    for t in times:
        if is_hourly:
            start = t
            end = t + pd.Timedelta(minutes=30)
        else:
            start = t
            end = t + pd.Timedelta(days=1)
        plan = calculate_overlap(df_period, start, end)
        plan_list.append({'slot_start': t, 'План': plan})
    return pd.DataFrame(plan_list).set_index('slot_start')

def calculate_forecast(df_forecast, times, start_dt=None, end_dt=None, is_hourly=True):
    """
    Рассчитывает прогноз по заданным временным слотам.

    Args:
        df_forecast (pd.DataFrame): DataFrame прогноза
        times (list[datetime]): Список временных слотов
        start_dt (datetime): Начало периода (для дневного режима)
        end_dt (datetime): Конец периода (для дневного режима)
        is_hourly (bool): Режим — по часам (True) или по дням (False)

    Returns:
        pd.DataFrame: DataFrame с колонкой 'Прогноз' (в часах)
    """
    if df_forecast.empty:
        return pd.DataFrame({'slot_start': times, 'Прогноз': [0.0] * len(times)}).set_index('slot_start')

    df_fc = df_forecast.copy()
    df_fc['ts'] = pd.to_datetime(df_fc['Дата'] + ' ' + df_fc['Время'])
    if not is_hourly:
        df_fc = df_fc[(df_fc['ts'] >= start_dt) & (df_fc['ts'] < end_dt)]

    forecast_list = []
    for t in times:
        if is_hourly:
            mask = df_fc['ts'] == t
        else:
            mask = (df_fc['ts'] >= t) & (df_fc['ts'] < t + pd.Timedelta(days=1))
        forecast_list.append({'slot_start': t, 'Прогноз': df_fc.loc[mask, 'Прогноз'].sum()})

    return pd.DataFrame(forecast_list).set_index('slot_start')

def finalize_slot_df(plan_df, forecast_df, year=None, month=None):
    """
    Объединяет план и прогноз, рассчитывает равномерность.

    Args:
        plan_df (pd.DataFrame): DataFrame с планом
        forecast_df (pd.DataFrame): DataFrame с прогнозом

    Returns:
        pd.DataFrame: Итоговый DataFrame с колонками:
            - 'slot_start' (datetime)
            - 'План' (float)
            - 'Прогноз' (float)
            - 'Равномерность' (float)
    """
    # Объединяем по слотам
    slot_df = pd.concat([plan_df, forecast_df], axis=1).fillna(0)
    slot_df = slot_df.reset_index()
    slot_df['slot_start'] = pd.to_datetime(slot_df['slot_start'])

    # 🔥 Фильтрация по году и месяцу, если переданы
    if year is not None and month is not None:
        slot_df = slot_df[
            (slot_df['slot_start'].dt.year == year) &
            (slot_df['slot_start'].dt.month == month)
        ]

    # Проверка на пустоту
    if slot_df.empty:
        st.warning("⚠️ Нет данных для расчёта равномерности.")
        return pd.DataFrame()

    # Рассчитываем только по данным выбранного периода
    total_plan = slot_df['План'].sum()
    total_forecast = slot_df['Прогноз'].sum()

    # Защита от деления на 0
    k = total_plan / total_forecast if total_forecast > 0 else 0.0

    slot_df['Равномерность'] = slot_df['Прогноз'] * k

    return slot_df


def prepare_slot_data(mode, df_filtered, df_forecast, min_date, max_date):
    """
    Подготавливает данные по временным слотам для отображения.

    Args:
        mode (str): Режим отображения ('По часам' или 'По дням')
        df_filtered (pd.DataFrame): Отфильтрованный DataFrame активности
        df_forecast (pd.DataFrame): DataFrame прогноза
        min_date (date): Минимальная доступная дата
        max_date (date): Максимальная доступная дата

    Returns:
        dict: Словарь с данными:
            - 'slot_df' (pd.DataFrame): Итоговые данные
            - 'selected_date' (date): Выбранная дата
            - 'year' (int): Выбранный год
            - 'month' (int): Выбранный месяц
            - 'start_dt' (datetime): Начало периода
            - 'end_dt' (datetime): Конец периода
    """
    selected_date = year = month = None
    start_dt = end_dt = None

    if mode == "По часам":
        selected_date = st.session_state.selected_date
        start_dt = pd.to_datetime(selected_date)
        end_dt = start_dt + pd.Timedelta(days=1)
    else:
        # Для режима "По дням" start_dt и end_dt остаются None
        start_dt = pd.to_datetime(min_date)
        end_dt = pd.to_datetime(max_date) + pd.Timedelta(days=1)

    # --- 🔥 ФИЛЬТРАЦИЯ ПО ПЕРИОДУ ---
    if start_dt is not None and end_dt is not None:
        df_period = filter_by_period(df_filtered, start_dt, end_dt)
    else:
        df_period = df_filtered.copy()

    # 3. Создание слотов и расчёт
    if mode == "По часам":
        times = pd.date_range(start=start_dt, end=end_dt, freq='30min', inclusive='left').tolist()
        plan_df = calculate_plan(df_period, times, is_hourly=True)
        forecast_df = calculate_forecast(df_forecast, times, is_hourly=True)
    else:
        # Для режима "По дням" используем диапазон от min_date до max_date
        times = pd.date_range(start=min_date, end=max_date, freq='D', inclusive='left').tolist()
        df_period = df_filtered.copy()  # Полный период

        # --- 🔥 ДОПОЛНИТЕЛЬНАЯ ФИЛЬТРАЦИЯ ПО МЕСЯЦУ ---
        if 'year' in st.session_state and 'month' in st.session_state:
            selected_year = st.session_state.selected_year
            selected_month = st.session_state.selected_month
            df_period = df_period[
                df_period['start'].dt.year == selected_year
                ]
            df_period = df_period[
                df_period['start'].dt.month == selected_month
                ]

        plan_df = calculate_plan(df_period, times, is_hourly=False)
        forecast_df = calculate_forecast(df_forecast, times, start_dt=start_dt, end_dt=end_dt, is_hourly=False)

    # 4. Объединение и финальный расчёт
    slot_df = finalize_slot_df(plan_df, forecast_df)

    return {
        'slot_df': slot_df,
        'selected_date': selected_date if mode == "По часам" else None,
        'year': year,
        'month': month,
        'start_dt': start_dt,
        'end_dt': end_dt
    }

def render_chart_and_table(slot_df, mode, selected_date=None, year=None, month=None):
    """
    Строит график и отображает таблицу на основе slot_df
    """
    if slot_df.empty:
        st.warning("⚠️ Нет данных для отображения.")
        return

    # --- 🔥 Фильтрация по месяцу в режиме "По дням" ---
    if mode == "По дням" and year and month:
        slot_df = slot_df[
            (slot_df['slot_start'].dt.year == year) &
            (slot_df['slot_start'].dt.month == month)
        ]

    # Если после фильтрации данные пустые
    if slot_df.empty:
        st.warning("⚠️ Нет данных для отображения в выбранном периоде.")
        return

    # --- Построение графика ---
    title = (
        f"Почасовой обзор — {selected_date}"
        if mode == "По часам"
        else f"Дневной обзор — {year}-{month:02d}"
    )
    fig = px.line(
        slot_df,
        x='slot_start',
        y=['План', 'Прогноз', 'Равномерность'],
        labels={'slot_start': 'Время/Дата', 'value': 'Человеко-часы'},
        title=title,
        color_discrete_map={'План': 'green', 'Прогноз': 'red', 'Равномерность': 'orange'},
    )
    fig.update_traces(visible='legendonly', selector=dict(name='Равномерность'))

    if mode == "По часам":
        fig.update_xaxes(tickformat='%H:%M', dtick=30 * 60 * 1000)
    else:
        fig.update_xaxes(tickformat='%Y-%m-%d', dtick="D1")

    # Установка нижней границы Y-оси в 0
    max_value = slot_df[['План', 'Прогноз', 'Равномерность']].max().max()
    fig.update_layout(yaxis=dict(range=[0, max_value + 1 if max_value > 0 else 1]))

    if slot_df['План'].sum() + slot_df['Прогноз'].sum() == 0:
        st.warning("Нет данных для построения графика. Проверьте:")
        st.markdown("- Корректность временных интервалов")
        st.markdown("- Наличие активностей в выбранный период")
    else:
        st.plotly_chart(fig, use_container_width=True)

    # --- Таблица данных ---
    st.subheader("📊 Таблица данных: План и Прогноз")

    display_df = slot_df[['slot_start', 'План', 'Прогноз']].copy()
    if mode == "По часам":
        display_df['Время'] = display_df['slot_start'].dt.strftime('%H:%M')
        cols = ['Время', 'План', 'Прогноз']
    else:
        display_df['Дата'] = display_df['slot_start'].dt.strftime('%Y-%m-%d')
        cols = ['Дата', 'План', 'Прогноз']
    display_df['Дельта'] = display_df['План'] - display_df['Прогноз']

    # --- Стилизация таблицы ---
    def color_delta(val):
        color = 'green' if val > 0 else 'red' if val < 0 else 'gray'
        return f'color: {color}'

    styled_df = display_df[cols + ['Дельта']].style \
        .format({'План': '{:.1f}', 'Прогноз': '{:.1f}', 'Дельта': '{:+.1f}'}) \
        .map(color_delta, subset='Дельта') \
        .set_properties(**{'text-align': 'center'}) \
        .set_table_styles([{
        'selector': 'th',
        'props': [('background-color', '#3498db'), ('color', 'white')]
    }])

    st.dataframe(
        styled_df,
        use_container_width=True,
        height=400,
        hide_index=True
    )

def render_export_buttons(slot_df, mode, df_filtered, df_forecast, selected_channels, selected_variants, selected_date=None, year=None, month=None):
    """
    Отображает кнопки экспорта данных в Excel.
    Поддерживает режимы:
    - 'По часам' → использует selected_date
    - 'По дням'  → использует year и month
    """
    # Подготовка данных для "По интервалам"
    output_df = slot_df.copy()
    output_df['Дельта'] = output_df['План'] - output_df['Прогноз']
    output_df['Дата'] = output_df['slot_start'].dt.date

    if mode == "По часам":
        output_df['Время'] = output_df['slot_start'].dt.strftime('%H:%M')
        cols_sel = ['Дата', 'Время', 'План', 'Прогноз', 'Дельта']
    else:
        cols_sel = ['Дата', 'План', 'Прогноз', 'Дельта']

    # Формирование Excel-файла "По интервалам"
    buf_sel = io.BytesIO()
    with pd.ExcelWriter(buf_sel, engine='openpyxl') as writer:
        output_df[cols_sel].to_excel(writer, sheet_name='Интервалы', index=False)
        ws = writer.sheets['Интервалы']
        for i, col in enumerate(cols_sel, start=1):
            w = max(output_df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(i)].width = w
    buf_sel.seek(0)

    # Формирование имени файла
    timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M')
    if mode == "По часам":
        file_label = f"отчет_{selected_date}_{timestamp}.xlsx"
    else:
        file_label = f"отчет_{year}-{month:02d}_{timestamp}.xlsx"

    # Рендер кнопок
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(
            label="⬇️ По выбранным интервалам",
            data=buf_sel,
            file_name=f"интервалы_{file_label}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # --- Кнопка экспорта KPI за месяц ---
    if mode == "По дням" and year and month:
        with c2:
            with st.spinner("⏳ Формируем KPI за месяц..."):
                kpi_df = calculate_monthly_kpi(df_filtered, df_forecast, selected_channels, selected_variants, year, month)
                buf_kpi = io.BytesIO()
                with pd.ExcelWriter(buf_kpi, engine='openpyxl') as writer:
                    kpi_df.to_excel(writer, sheet_name='KPI_30м', index=False)
                    ws_kpi = writer.sheets['KPI_30м']
                    for i, col in enumerate(kpi_df.columns, start=1):
                        w = max(kpi_df[col].astype(str).map(len).max(), len(col)) + 2
                        ws_kpi.column_dimensions[get_column_letter(i)].width = w
                buf_kpi.seek(0)

                st.download_button(
                    label="⬇️ По выбранным интервалам за месяц (30-минутные слоты)",
                    data=buf_kpi,
                    file_name=f"KPI_{year}-{month:02d}_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

@st.cache_data(hash_funcs={pd.DataFrame: hash_dataframe})
def calculate_monthly_kpi(df_filtered, df_forecast, selected_channels, selected_variants, year, month):
    start_month = pd.Timestamp(year=year, month=month, day=1)
    end_month = start_month + pd.offsets.MonthEnd(1) + pd.Timedelta(days=1)
    slots_month = pd.date_range(start=start_month, end=end_month, freq='30min', inclusive='left')

    df_act_m = df_filtered[(df_filtered['start'] < end_month) & (df_filtered['end'] > start_month)]

    if not df_forecast.empty:
        df_fc_m = df_forecast.copy()
        df_fc_m['ts'] = pd.to_datetime(df_fc_m['Дата'] + ' ' + df_fc_m['Время'])
        df_fc_m = df_fc_m[(df_fc_m['ts'] >= start_month) & (df_fc_m['ts'] < end_month)]

        if 'skill_variant' not in df_fc_m.columns:
            df_fc_m['skill_variant'] = 'Не задано'
    else:
        df_fc_m = pd.DataFrame(columns=['ts', 'Прогноз', 'Канал коммуникации', 'skill_variant'])

    kpi_data = []

    for channel in selected_channels:
        df_act_channel = df_act_m[df_act_m['main_act'].isin(CHANNEL_MAPPING[channel])]

        for skill_variant in selected_variants:
            # 🔁 Получаем системную группу из VARIANT_TO_SYSTEM
            system_group = VARIANT_TO_SYSTEM.get(skill_variant, skill_variant)

            # Фильтруем активность по оригинальной skill_variant
            df_act_skill = df_act_channel[df_act_channel['skill_variant'] == skill_variant]

            # 🔁 Фильтруем прогноз по системной группе
            df_fc_channel = df_fc_m[
                (df_fc_m['Канал коммуникации'] == channel) &
                (df_fc_m['system_group'] == system_group)
            ]

            for ts in slots_month:
                plan = calculate_overlap(df_act_skill, ts, ts + pd.Timedelta(minutes=30))
                forecast = df_fc_channel[df_fc_channel['ts'] == ts]['Прогноз'].sum()
                kpi_data.append({
                    'Дата': ts.date(),
                    'Время': ts.time(),
                    'Канал коммуникации': channel,
                    'Скилл группа': skill_variant,
                    'План': plan,
                    'Прогноз': forecast
                })

    kpi_df = pd.DataFrame(kpi_data)
    kpi_df['Дельта'] = kpi_df['План'] - kpi_df['Прогноз']
    return kpi_df

@st.cache_data
def load_activity(file) -> pd.DataFrame:
    """
    Загружает и валидирует файл активности.

    Args:
        file (io.BytesIO): Загруженный XLSX-файл

    Returns:
        pd.DataFrame: Обработанный DataFrame активности
    """
    try:
        df = pd.read_excel(file, dtype={
            'activity_date': str,
            'start_time': str,
            'end_time': str,
            'main_act': str,
            'Скилл-группа': str
        })

        # Проверка обязательных колонок
        required_columns = ['activity_date', 'start_time', 'end_time', 'Скилл-группа', 'main_act']
        if not all(col in df.columns for col in required_columns):
            st.error(f"Отсутствуют колонки: {set(required_columns) - set(df.columns)}")
            return pd.DataFrame()

        # Сохраняем оригинал в skill_variant и мапим в system_group
        df['skill_variant'] = df['Скилл-группа']
        df['system_group'] = df['skill_variant'] \
            .map(VARIANT_TO_SYSTEM) \
            .fillna(df['skill_variant'])
            # Создание временных меток
        df['start'] = pd.to_datetime(
            df['activity_date'] + ' ' + df['start_time'],
            errors='coerce'
        )
        df['end'] = pd.to_datetime(
            df['activity_date'] + ' ' + df['end_time'],
            errors='coerce'
        )

        # Проверка ошибок конвертации
        if df[['start', 'end']].isnull().any().any():
            st.error("Ошибка в формате даты/времени. Проверьте входные данные.")
            return pd.DataFrame()

        # Корректировка перехода через полночь
        df.loc[df['end'] <= df['start'], 'end'] += pd.Timedelta(days=1)

        return df.dropna(subset=['start', 'end'])

    except Exception as e:
        st.error(f"Ошибка загрузки активности: {str(e)}")
        return pd.DataFrame()


@st.cache_data
def load_forecast(file) -> pd.DataFrame:
    """
    Загружает и валидирует файл прогноза.

    Args:
        file (io.BytesIO | list[io.BytesIO]): Загруженный(е) XLSX-файл(ы)

    Returns:
        pd.DataFrame: Обработанный DataFrame прогноза
    """
    if file is None:
        return pd.DataFrame()

    if isinstance(file, list):
        if not file:  # Если список пуст — возвращаем пустой DataFrame
            return pd.DataFrame()

        dfs = []
        for f in file:
            df = load_forecast_single(f)
            if not df.empty:
                dfs.append(df)
        return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

    return load_forecast_single(file)


def load_forecast_single(file) -> pd.DataFrame:
    """
    Вспомогательная функция для загрузки одного файла прогноза
    """
    try:
        df = pd.read_excel(file, dtype={
            'Дата': str,
            'Время': str,
            'Скилл группа': str,
            'Канал коммуникации': str,
            'Прогноз Raw': float,
            'Прогноз Abs_new': float,
            'Прогноз Full': float
        }).rename(columns={'Скилл группа': 'skill_variant'})

        required = ['skill_variant', 'Канал коммуникации', 'Прогноз Raw', 'Прогноз Abs_new', 'Прогноз Full']
        if not all(col in df.columns for col in required):
            st.error(f"В прогнозе отсутствуют колонки: {set(required) - set(df.columns)}")
            return pd.DataFrame()

        df['system_group'] = df['skill_variant'] \
            .map(VARIANT_TO_SYSTEM) \
            .fillna(df['skill_variant'])

        return df

    except Exception as e:
        st.error(f"Ошибка загрузки прогноза: {str(e)}")
        return pd.DataFrame()


def calculate_overlap(df_activity, slot_start, slot_end):
    """
    Рассчитывает пересечение временных интервалов активности с заданным слотом.

    Args:
        df_activity (pd.DataFrame): DataFrame активности
        slot_start (datetime): Начало временного слота
        slot_end (datetime): Конец временного слота

    Returns:
        float: Время пересечения в часах
    """
    df = df_activity.copy()
    mask = (df['start'] < slot_end) & (df['end'] > slot_start)
    filtered = df[mask].copy()

    if filtered.empty:
        return 0.0

    filtered['overlap_start'] = filtered[['start']].apply(lambda x: max(x['start'], slot_start), axis=1)
    filtered['overlap_end'] = filtered[['end']].apply(lambda x: min(x['end'], slot_end), axis=1)

    total_seconds = (filtered['overlap_end'] - filtered['overlap_start']).dt.total_seconds().sum()
    return total_seconds / 3600

def load_and_validate_files():
    """
    Загружает файлы активности и прогноза из интерфейса Streamlit.

    Позволяет пользователю загрузить XLSX-файлы через боковую панель.
    Проверяет, загружен ли файл активности. Если нет — выводит инструкцию.

    Returns:
        tuple[pd.DataFrame, io.BytesIO, io.BytesIO] | tuple[None, None, None]:
            Возвращает кортеж из:
            - DataFrame активности (pd.DataFrame)
            - activity_file (io.BytesIO): файл активности
            - forecast_file (io.BytesIO): файл прогноза
            Если файл активности не загружен — возвращает (None, None, None)
    """
    with st.sidebar.expander("⚙️ **Загрузка данных**", expanded=True):
        activity_files = st.file_uploader(
            "Файлы активности (можно выбрать несколько)",
            type="xlsx",
            help="Загрузите один или несколько XLSX файлов с данными активности",
            key="activity_uploader",
            accept_multiple_files=True
        )

        forecast_files = st.file_uploader(
            "Файлы прогноза (можно выбрать несколько)",
            type="xlsx",
            help="Загрузите один или несколько XLSX файлов с прогнозами",
            key="forecast_uploader",
            accept_multiple_files=True
        )

        if activity_files:
            for i, f in enumerate(activity_files):
                st.success(f"✅ Файл активности {i + 1}: {f.name}")
        if forecast_files:
            for i, f in enumerate(forecast_files):
                st.success(f"✅ Файл прогноза {i + 1}: {f.name}")

    if not activity_files:
        st.info("""
              ## 🚀 Инструкция
              1. Скачайте подробное расписание из TWFM, предварительно добавьте выгрузку скилл-группы и загрузите файл(ы) активности через боковую панель. 
              2. Загрузите файл(ы) прогноза
              3. Настройте фильтры для анализа
              4. Используйте данные через интерактивные визуализации
              """)
        return None, None, None

    with st.spinner('🌀 **Обработка данных...**'):
        # Загрузка и объединение всех файлов активности
        dfs_act = []
        for file in activity_files:
            df = load_activity(file)
            if not df.empty:
                dfs_act.append(df)
            else:
                st.warning(f"⚠️ Ошибка загрузки файла активности: {file.name}")

        if not dfs_act:
            st.error("⚠️ Ни один файл активности не загружен корректно.")
            return None, None, None

        df_act_combined = pd.concat(dfs_act, ignore_index=True)

    return df_act_combined, activity_files, forecast_files

def apply_filters_and_process(df_act, filters, forecast_file):
    """
    Применяет фильтры к данным активности и обрабатывает прогноз.

    Фильтрует данные по выбранным вариантам и каналам.
    Если активности по выбранным каналам нет — выводит предупреждение.
    Если фильтры не дают данных — выводит ошибку.

    Args:
        df_act (pd.DataFrame): DataFrame активности
        filters (dict): Словарь с фильтрами, возвращаемый из `render_filters`
        forecast_file (io.BytesIO): файл прогноза

    Returns:
        tuple[pd.DataFrame, pd.DataFrame] | tuple[None, None]:
            - Отфильтрованный DataFrame активности
            - DataFrame прогноза
            Если фильтры не дают данных — возвращает (None, None)
    """
    selected_variants = filters['selected_variants']
    selected_channels = filters['selected_channels']
    selected_main_acts = filters['selected_main_acts']
    selected_forecast_col = filters['selected_forecast_col']

    selected_system_groups = list({VARIANT_TO_SYSTEM.get(v, v) for v in selected_variants})
    df_filtered = filter_data(df_act, selected_variants, selected_main_acts)
    df_forecast = process_forecast(forecast_file, selected_system_groups, selected_channels, selected_forecast_col)

    if df_filtered.empty and selected_channels:
        st.warning("⚠️ Нет активности по выбранным каналам коммуникации.")
        st.markdown("Попробуйте выбрать другие каналы или измените период анализа.")

    if df_filtered.empty:
        reason = "при выбранных фильтрах" if selected_variants or selected_channels else "в исходных данных"
        st.error(f"Нет данных для отображения {reason}. Проверьте:")
        st.markdown("- Выбранные скилл-группы")
        st.markdown("- Выбранные каналы коммуникации")
        return None, None

    return df_filtered, df_forecast


def get_period_params(df_filtered):
    """
    Возвращает параметры выбранного периода: режим, дату/месяц, временные границы.

    Args:
        df_filtered (pd.DataFrame): Отфильтрованный DataFrame активности

    Returns:
        dict: {
            'mode': str,                     # 'По часам' / 'По дням'
            'selected_date': date,           # выбранная дата (если режим 'По часам')
            'year': int,                     # выбранный год (если режим 'По дням')
            'month': int,                    # выбранный месяц (если режим 'По дням')
            'start_dt': pd.Timestamp,        # начало периода
            'end_dt': pd.Timestamp,          # конец периода
            'min_date': date,                # минимальная доступная дата
            'max_date': date,                # максимальная доступная дата
        }
    """
    if df_filtered.empty:
        st.error("⚠️ Нет данных для обработки даты и режима.")
        return {}

        # Определяем общий диапазон дат один раз
    min_date = df_filtered['start'].min().date()
    max_date = df_filtered['start'].max().date()

    # Инициализируем session_state, если не задано
    if 'selected_date' not in st.session_state:
        st.session_state.selected_date = min_date

    if 'selected_year' not in st.session_state:
        st.session_state.selected_year = min_date.year

    if 'selected_month' not in st.session_state:
        st.session_state.selected_month = min_date.month

    # Режим отображения
    mode = st.sidebar.radio("Режим просмотра", ["По часам", "По дням"])

    selected_date = year = month = None
    start_dt = end_dt = None

    if mode == "По часам":
        current_date = st.session_state.selected_date

        # Автоматически корректируем дату, если вне диапазона
        if current_date < min_date:
            current_date = min_date
            st.session_state.selected_date = current_date
            st.warning("⚠️ Выбранная дата до начала данных. Установлена минимальная дата.")
        elif current_date > max_date:
            current_date = max_date
            st.session_state.selected_date = current_date
            st.warning("⚠️ Выбранная дата после окончания данных. Установлена максимальная дата.")

        selected_date = st.sidebar.date_input(
            "Выберите дату",
            value=current_date,
            min_value=min_date,
            max_value=max_date,
            key='period_selection_date_input'
        )

        if selected_date != st.session_state.selected_date:
            st.session_state.selected_date = selected_date
            st.rerun()

        start_dt = pd.to_datetime(selected_date)
        end_dt = start_dt + pd.Timedelta(days=1)

    else:
        # Получаем доступные годы
        available_years = sorted(df_filtered['start'].dt.year.unique())

        # Получаем доступные месяцы для выбранного года
        selected_year = st.session_state.selected_year
        available_months = sorted(
            df_filtered[df_filtered['start'].dt.year == selected_year]['start'].dt.month.unique()
        )

        # Выбор года
        year = st.sidebar.selectbox(
            "Год",
            options=available_years,
            index=available_years.index(selected_year) if selected_year in available_years else 0,
            key='period_selection_year'
        )

        # Обновляем доступные месяцы для нового года
        available_months = sorted(
            df_filtered[df_filtered['start'].dt.year == year]['start'].dt.month.unique()
        )

        # Если текущий месяц не в списке — выбираем первый
        selected_month = st.session_state.selected_month
        if selected_month not in available_months:
            selected_month = available_months[0]

        # Выбор месяца
        month = st.sidebar.selectbox(
            "Месяц",
            options=available_months,
            index=available_months.index(selected_month),
            key='period_selection_month'
        )

        # Сохраняем выбор в session_state
        st.session_state.selected_year = year
        st.session_state.selected_month = month

        start_dt = pd.Timestamp(year=year, month=month, day=1)
        end_dt = (start_dt + pd.offsets.MonthEnd(1)) + pd.Timedelta(days=1)

    return {
        'mode': mode,
        'selected_date': selected_date,
        'year': year,
        'month': month,
        'start_dt': start_dt,
        'end_dt': end_dt,
        'min_date': min_date,
        'max_date': max_date
    }

def render_results(slot_df, mode, selected_date, year, month, df_filtered, df_forecast, selected_channels, selected_variants):
    """
    Отображает результаты в зависимости от режима (по часам или по дням).
    """
    if mode == "По часам":
        if selected_date is not None and not slot_df.empty:
            render_chart_and_table(slot_df, mode, selected_date=selected_date)
            render_export_buttons(slot_df, mode, df_filtered, df_forecast, selected_channels, selected_variants, selected_date=selected_date)
        else:
            st.warning("⚠️ Нет данных для отображения в выбранной дате.")
    else:
        if year is not None and month is not None and not slot_df.empty:
            render_chart_and_table(slot_df, mode, year=year, month=month)
            render_export_buttons(slot_df, mode, df_filtered, df_forecast, selected_channels, selected_variants, year=year, month=month)
        else:
            st.warning("⚠️ Нет данных для отображения в выбранном периоде.")

def configure_page():
    """
    Настраивает параметры страницы Streamlit.
    Устанавливает макет, заголовок, иконку и состояние боковой панели.
    """
    st.set_page_config(
        layout="wide",
        page_title="Анализ активности",
        page_icon="📊",
        initial_sidebar_state="expanded"
    )


def run_streamlit():
    configure_page()
    inject_custom_css()
    st.title("📈 Анализ активности и прогноза")

    # 1. Загрузка и валидация файлов
    result = load_and_validate_files()
    df_act_combined, activity_file, forecast_file = result

    if df_act_combined is None:
        return

    # 2. Применение фильтров
    filters = render_filters(df_act_combined)
    selected_channels = filters['selected_channels']
    selected_variants = filters['selected_variants']
    selected_main_acts = filters['selected_main_acts']
    selected_forecast_col = filters['selected_forecast_col']

    df_filtered, df_forecast = apply_filters_and_process(df_act_combined, filters, forecast_file)
    if df_filtered is None:
        return

    # 3. Обработка даты и режима
    period_data = get_period_params(df_filtered)
    if not period_data:
        return

    mode = period_data['mode']
    selected_date = period_data['selected_date']
    year = period_data['year']
    month = period_data['month']
    start_dt = period_data['start_dt']
    end_dt = period_data['end_dt']

    # 4. Подготовка данных
    data = prepare_slot_data(
        mode,
        df_filtered,
        df_forecast,
        period_data['min_date'],
        period_data['max_date']
    )
    slot_df = data['slot_df']

    # 5. Визуализация и экспорт
    render_results(
        slot_df, mode, selected_date, year, month,
        df_filtered, df_forecast,
        selected_channels, selected_variants
    )


if __name__ == "__main__":
    run_streamlit()
